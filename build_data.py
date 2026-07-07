import os
import glob
import json
import re
from openpyxl import load_workbook

MONTH_ORDER = [
    'Yanvar', 'Fevral', 'Mart', 'Aprel', 'May', 'Iyun',
    'Iyul', 'Avgust', 'Sentabr', 'Oktabr', 'Noyabr', 'Dekabr'
]

POLLUTANT_MAP = {
    'SO2': 'SO2',
    'NO2': 'NO2',
    'NH3': 'NH3',
    'HF': 'HF',
    'NO': 'NO',
    'Fenol': 'Fenol',
    'CO': 'CO',
    'CL': 'CL',
    'Chang': 'Chang'
}

MEASURE_MAP = {
    'OСrta': 'Mean',
    "O'rtacha": 'Mean',
    'O`rtacha': 'Mean',
    'Maksimal': 'Max',
    'Maks.': 'Max',
    'Mean': 'Mean',
    'Max': 'Max'
}

MONTH_TRANSLATIONS = {
    'январ': 'Yanvar',
    'феврал': 'Fevral',
    'март': 'Mart',
    'апрел': 'Aprel',
    'май': 'May',
    'июн': 'Iyun',
    'июл': 'Iyul',
    'август': 'Avgust',
    'сентябр': 'Sentabr',
    'октябр': 'Oktabr',
    'ноябр': 'Noyabr',
    'декабр': 'Dekabr'
}


def normalize_month(value):
    if not value:
        return None
    s = str(value).strip()
    lower = s.lower()
    for month in MONTH_ORDER:
        if lower.startswith(month.lower()):
            return month
    for key, month in MONTH_TRANSLATIONS.items():
        if key in lower:
            return month
    return s.capitalize()


def normalize_pollutant(name):
    if not name:
        return None
    s = str(name).strip()
    if s.lower() == 'none' or not s:
        return None
    # Check for key matching
    for key in POLLUTANT_MAP:
        if key in s:
            return POLLUTANT_MAP[key]
    # Return None unless it's a known pollutant
    return None


def normalize_measure(value):
    if not value:
        return None
    s = str(value).strip()
    if not s or s.lower() == 'none':
        return None
    if s in MEASURE_MAP:
        return MEASURE_MAP[s]
    # Check common patterns
    lower = s.lower()
    if lower.startswith('o') or 'mean' in lower or 'average' in lower:
        return 'Mean'
    if lower.startswith('m') or 'max' in lower or 'maks' in lower:
        return 'Max'
    return None


def parse_pollutant_file(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        return None

    header_row = rows[0]
    second_row = rows[1] if len(rows) > 1 else None
    columns = []

    # Detect two-header format by checking second row for measure keywords
    has_measures_row = False
    if second_row:
        measure_keywords = ['rta', 'maksimal', 'max', 'mean', 'maks']
        for cell in second_row:
            if cell:
                cell_str = str(cell).lower()
                if any(kw in cell_str for kw in measure_keywords):
                    has_measures_row = True
                    break

    if has_measures_row:
        # Two-header format: compound headers + measures
        # Handle pattern where alternating columns have None headers with measures
        last_pollutant = None
        for idx in range(len(header_row)):
            if idx == 0:
                columns.append({'type': 'month'})
                continue
            
            header_val = header_row[idx]
            measure_val = second_row[idx] if idx < len(second_row) else None
            
            header_str = str(header_val).strip() if header_val else ''
            measure_str = str(measure_val).strip() if measure_val else ''
            
            # Handle: if header is None but measure exists and last was pollutant
            if (not header_str or header_str.lower() == 'none') and measure_str and last_pollutant:
                measure = normalize_measure(measure_str)
                if measure:
                    columns.append({'type': 'pollutant', 'pollutant': last_pollutant, 'measure': measure})
                else:
                    columns.append({'type': 'ignore'})
                continue
            
            # Normal case: header has pollutant name
            if not header_str or header_str.lower() == 'none':
                columns.append({'type': 'ignore'})
                last_pollutant = None
                continue
            
            pollutant = normalize_pollutant(header_str)
            measure = normalize_measure(measure_str)
            
            if pollutant and measure:
                columns.append({'type': 'pollutant', 'pollutant': pollutant, 'measure': measure})
                last_pollutant = pollutant
            else:
                columns.append({'type': 'ignore'})
                last_pollutant = None
    else:
        # Single-header format or underscore-separated
        for idx, label in enumerate(header_row):
            if idx == 0:
                columns.append({'type': 'month'})
                continue
            if not label:
                columns.append({'type': 'ignore'})
                continue
            text = str(label).strip()
            if text.lower() == 'none':
                columns.append({'type': 'ignore'})
                continue
            if '_' in text:
                parts = text.split('_', 1)
                pollutant = normalize_pollutant(parts[0])
                measure = normalize_measure(parts[1]) if len(parts) > 1 else 'Mean'
                if pollutant and measure:
                    columns.append({'type': 'pollutant', 'pollutant': pollutant, 'measure': measure})
                else:
                    columns.append({'type': 'ignore'})
            else:
                columns.append({'type': 'ignore'})

    # Parse data rows
    data = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        month = normalize_month(row[0])
        if month not in MONTH_ORDER:
            continue
        data[month] = {}
        
        for idx, cell in enumerate(row):
            if idx >= len(columns):
                break
            col = columns[idx]
            if col['type'] != 'pollutant':
                continue
            
            pollutant = col['pollutant']
            measure = col['measure']
            
            if not pollutant or not measure:
                continue
            
            if pollutant not in data[month]:
                data[month][pollutant] = {}
            
            value = None if cell is None else cell
            data[month][pollutant][measure] = value

    # Build final structure with all months
    normalized = {}
    for month in MONTH_ORDER:
        normalized[month] = data.get(month, {})
    return normalized


def parse_wind_file(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)]
    wind_years = {}
    directions = []
    current_year = None
    header_directions = None

    for row in rows:
        if not row or not row[0]:
            continue
        first = str(row[0]).strip()
        
        # Check for year markers
        if 'shamol' in first.lower() and 'yil' in first.lower():
            year_match = re.search(r'(\d{4})', first)
            if year_match:
                current_year = year_match.group(1)
                wind_years[current_year] = {}
                header_directions = None
            continue

        if current_year is None:
            continue
        
        # Skip header rows
        if header_directions is None and ('shamol' in first.lower() or 'oylar' in first.lower()):
            header_directions = [str(cell).strip() for cell in row[1:] if cell]
            if not directions:
                directions = header_directions
            continue

        month = normalize_month(row[0])
        if month not in MONTH_ORDER:
            continue

        if header_directions:
            values = {}
            for idx, direction in enumerate(header_directions):
                col_idx = idx + 1
                value = row[col_idx] if col_idx < len(row) else None
                values[direction] = value
            wind_years[current_year][month] = values

    # Fill in missing months
    for year in wind_years:
        for month in MONTH_ORDER:
            if month not in wind_years[year]:
                wind_years[year][month] = {d: None for d in directions}

    return wind_years, directions


def build_data():
    workbook_files = sorted(glob.glob('*.xlsx'))
    final = {
        'months': MONTH_ORDER,
        'years': {},
        'pollutants': list(POLLUTANT_MAP.values()),
        'wind': {
            'years': {},
            'directions': []
        }
    }

    for workbook in workbook_files:
        print('Parsing', workbook)
        basename = os.path.basename(workbook).lower()
        if 'shamol' in basename:
            wind_years, directions = parse_wind_file(workbook)
            final['wind']['years'] = wind_years
            final['wind']['directions'] = directions
        elif 'gidrom' in basename:
            # Skip climate/weather files
            continue
        else:
            year = os.path.splitext(os.path.basename(workbook))[0]
            normalized = parse_pollutant_file(workbook)
            if normalized:
                final['years'][year] = normalized

    with open(os.path.join('data', 'samarqand_data.json'), 'w', encoding='utf-8') as f:
        json.dump(final, f, ensure_ascii=False, indent=2)
    print('Wrote data/samarqand_data.json')


if __name__ == '__main__':
    build_data()
